#!/usr/bin/env python3
"""
Excel表格合并插件
将两个Excel文件的指定列合并到一个新Excel文件中
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def get_desktop_path():
    """获取桌面路径"""
    if 'USERPROFILE' in os.environ:
        desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
        if os.path.exists(desktop):
            return desktop
    return None


def read_excel_column(file_path, sheet_name, column_index, start_row=1):
    """
    读取Excel文件中的指定列

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称(默认第一个工作表)
        column_index: 列索引(从1开始)
        start_row: 起始行(从1开始)

    Returns:
        list: 列数据列表
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active if sheet_name is None else wb[sheet_name]

        data = []
        for row in ws.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index, values_only=True):
            data.append(row[0] if row[0] is not None else '')

        wb.close()
        return data
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        return None


def create_merged_excel(file1_path, file2_path, output_path=None,
                       file1_column=1, file2_column=2,
                       file1_sheet=None, file2_sheet=None,
                       output_filename="merged_excel.xlsx"):
    """
    合并两个Excel文件的指定列

    Args:
        file1_path: 第一个Excel文件路径
        file2_path: 第二个Excel文件路径
        output_path: 输出文件路径(默认桌面)
        file1_column: 第一个文件的列索引(从1开始)
        file2_column: 第二个文件的列索引(从1开始)
        file1_sheet: 第一个文件的工作表名(可选)
        file2_sheet: 第二个文件的工作表名(可选)
        output_filename: 输出文件名

    Returns:
        str: 输出文件路径,失败返回None
    """
    print("=" * 60)
    print("📊 Excel表格合并工具")
    print("=" * 60)
    print()

    # 验证输入文件
    if not os.path.exists(file1_path):
        print(f"❌ 错误: 第一个文件不存在: {file1_path}")
        return None

    if not os.path.exists(file2_path):
        print(f"❌ 错误: 第二个文件不存在: {file2_path}")
        return None

    print(f"📁 文件1: {os.path.basename(file1_path)}")
    print(f"   列: 第{file1_column}列")
    print()

    print(f"📁 文件2: {os.path.basename(file2_path)}")
    print(f"   列: 第{file2_column}列")
    print()

    # 读取数据
    print("⏳ 正在读取数据...")
    data1 = read_excel_column(file1_path, file1_sheet, file1_column)
    data2 = read_excel_column(file2_path, file2_sheet, file2_column)

    if data1 is None or data2 is None:
        return None

    print(f"✅ 文件1读取完成,共 {len(data1)} 行数据")
    print(f"✅ 文件2读取完成,共 {len(data2)} 行数据")
    print()

    # 确定输出路径
    if output_path is None:
        output_path = get_desktop_path()

    if output_path is None:
        print("❌ 错误: 无法获取桌面路径")
        return None

    output_file = os.path.join(output_path, output_filename)

    # 创建新的Excel文件
    print("⏳ 正在合并数据...")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "合并结果"

        # 设置表头
        ws.cell(row=1, column=1, value="文件1_第1列")
        ws.cell(row=1, column=2, value="文件2_第2列")

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment

        ws.cell(row=1, column=2).font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = header_alignment

        # 合并数据
        max_rows = max(len(data1), len(data2))

        for i in range(max_rows):
            row_num = i + 2  # 从第2行开始(第1行是表头)

            # 写入文件1的数据
            if i < len(data1):
                value1 = data1[i]
            else:
                value1 = ''

            # 写入文件2的数据
            if i < len(data2):
                value2 = data2[i]
            else:
                value2 = ''

            ws.cell(row=row_num, column=1, value=value1)
            ws.cell(row=row_num, column=2, value=value2)

        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30

        # 保存文件
        wb.save(output_file)
        wb.close()

        print("✅ 数据合并完成!")
        print()
        print("=" * 60)
        print(f"📄 输出文件: {output_file}")
        print(f"📊 合并行数: {max_rows}")
        print("=" * 60)
        print()
        print("✅ 操作成功完成!")

        return output_file

    except Exception as e:
        print(f"❌ 保存Excel文件失败: {e}")
        return None


def main():
    """主函数"""
    print("Excel表格合并工具")
    print("=" * 60)
    print()

    # 检查命令行参数
    if len(sys.argv) < 3:
        print("❌ 使用方法:")
        print()
        print("方法1: 合并两个Excel文件")
        print(f"  python {sys.argv[0]} <文件1路径> <文件2路径> [列1] [列2] [输出文件名]")
        print()
        print("参数说明:")
        print("  文件1路径: 第一个Excel文件的完整路径")
        print("  文件2路径: 第二个Excel文件的完整路径")
        print("  列1: (可选)第一个文件的列号,默认为1")
        print("  列2: (可选)第二个文件的列号,默认为2")
        print("  输出文件名: (可选)输出的Excel文件名,默认为merged_excel.xlsx")
        print()
        print("示例:")
        print(f"  python {sys.argv[0]} \"C:\\Users\\Desktop\\data1.xlsx\" \"C:\\Users\\Desktop\\data2.xlsx\"")
        print(f"  python {sys.argv[0]} data1.xlsx data2.xlsx 1 3 result.xlsx")
        print()
        sys.exit(1)

    # 获取参数
    file1 = sys.argv[1]
    file2 = sys.argv[2]
    col1 = int(sys.argv[3]) if len(sys.argv) > 3 else 1
    col2 = int(sys.argv[4]) if len(sys.argv) > 4 else 2
    output_name = sys.argv[5] if len(sys.argv) > 5 else "merged_excel.xlsx"

    # 执行合并
    result = create_merged_excel(
        file1_path=file1,
        file2_path=file2,
        output_path=None,  # 默认桌面
        file1_column=col1,
        file2_column=col2,
        output_filename=output_name
    )

    if result:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
